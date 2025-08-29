from django.shortcuts import render, redirect, get_object_or_404
from .models import Customer, InventoryItem, Order, Requirement, Payment, Supplier, Purchase, Notification, CustomerUser
from .forms import CustomerForm, InventoryItemForm, OrderForm, RequirementForm, PaymentForm, SupplierForm, PurchaseForm, CustomerUserRegistrationForm
from django.db.models import Sum, Q
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.forms import UserCreationForm
from django.contrib import messages
import csv
from django.http import HttpResponse
import io
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.db.models.functions import TruncMonth
from django.db.models import Count
from django.apps import apps
from rest_framework import viewsets, permissions, routers
from rest_framework.authtoken.models import Token
from rest_framework.authentication import TokenAuthentication
from rest_framework.decorators import action, api_view, authentication_classes, permission_classes
from rest_framework.response import Response
from django.http import FileResponse
import qrcode
import io
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from django.core.mail import send_mail
from django.utils import timezone
from django.contrib.auth import authenticate, login
from icalendar import Calendar, Event
from django.contrib.auth.views import LoginView
from django.utils.safestring import mark_safe

# API Serializers
from rest_framework import serializers
class CustomerSerializer(serializers.ModelSerializer):
    class Meta:
        model = Customer
        fields = '__all__'
class InventoryItemSerializer(serializers.ModelSerializer):
    class Meta:
        model = InventoryItem
        fields = '__all__'
class OrderSerializer(serializers.ModelSerializer):
    class Meta:
        model = Order
        fields = '__all__'
class RequirementSerializer(serializers.ModelSerializer):
    class Meta:
        model = Requirement
        fields = '__all__'
class PaymentSerializer(serializers.ModelSerializer):
    class Meta:
        model = Payment
        fields = '__all__'
class SupplierSerializer(serializers.ModelSerializer):
    class Meta:
        model = Supplier
        fields = '__all__'
class PurchaseSerializer(serializers.ModelSerializer):
    class Meta:
        model = Purchase
        fields = '__all__'

# API ViewSets
class CustomerViewSet(viewsets.ModelViewSet):
    queryset = Customer.objects.all()
    serializer_class = CustomerSerializer
    authentication_classes = [TokenAuthentication]
    permission_classes = [permissions.IsAuthenticated]
class InventoryItemViewSet(viewsets.ModelViewSet):
    queryset = InventoryItem.objects.all()
    serializer_class = InventoryItemSerializer
    authentication_classes = [TokenAuthentication]
    permission_classes = [permissions.IsAuthenticated]
class OrderViewSet(viewsets.ModelViewSet):
    queryset = Order.objects.all()
    serializer_class = OrderSerializer
    authentication_classes = [TokenAuthentication]
    permission_classes = [permissions.IsAuthenticated]
class RequirementViewSet(viewsets.ModelViewSet):
    queryset = Requirement.objects.all()
    serializer_class = RequirementSerializer
    authentication_classes = [TokenAuthentication]
    permission_classes = [permissions.IsAuthenticated]
class PaymentViewSet(viewsets.ModelViewSet):
    queryset = Payment.objects.all()
    serializer_class = PaymentSerializer
    authentication_classes = [TokenAuthentication]
    permission_classes = [permissions.IsAuthenticated]
class SupplierViewSet(viewsets.ModelViewSet):
    queryset = Supplier.objects.all()
    serializer_class = SupplierSerializer
    authentication_classes = [TokenAuthentication]
    permission_classes = [permissions.IsAuthenticated]
class PurchaseViewSet(viewsets.ModelViewSet):
    queryset = Purchase.objects.all()
    serializer_class = PurchaseSerializer
    authentication_classes = [TokenAuthentication]
    permission_classes = [permissions.IsAuthenticated]

# API Router
api_router = routers.DefaultRouter()
api_router.register(r'customers', CustomerViewSet)
api_router.register(r'inventory', InventoryItemViewSet)
api_router.register(r'orders', OrderViewSet)
api_router.register(r'requirements', RequirementViewSet)
api_router.register(r'payments', PaymentViewSet)
api_router.register(r'suppliers', SupplierViewSet)
api_router.register(r'purchases', PurchaseViewSet)

# Home view

def home(request):
    return render(request, 'core/home.html')

# Customers CRUD

@login_required
def customers(request):
    query = request.GET.get('q', '')
    customers = Customer.objects.all()
    if query:
        customers = customers.filter(
            Q(name__icontains=query) |
            Q(contact__icontains=query) |
            Q(address__icontains=query)
        )
    form = CustomerForm()
    if request.method == 'POST':
        form = CustomerForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('customers')
    return render(request, 'core/customers.html', {'customers': customers, 'form': form, 'query': query})

@login_required
def edit_customer(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    if request.method == 'POST':
        form = CustomerForm(request.POST, instance=customer)
        if form.is_valid():
            form.save()
            return redirect('customers')
    else:
        form = CustomerForm(instance=customer)
    return render(request, 'core/edit_customer.html', {'form': form, 'customer': customer})

@permission_required('core.delete_customer', raise_exception=True)
def delete_customer(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    if request.method == 'POST':
        customer.delete()
        return redirect('customers')
    return render(request, 'core/delete_customer.html', {'customer': customer})

@login_required
def customers_export(request):
    query = request.GET.get('q', '')
    customers = Customer.objects.all()
    if query:
        customers = customers.filter(
            Q(name__icontains=query) |
            Q(contact__icontains=query) |
            Q(address__icontains=query)
        )
    all_columns = [
        ('name', 'Name'),
        ('contact', 'Contact'),
        ('address', 'Address'),
        ('created_at', 'Created At'),
    ]
    selected = request.GET.getlist('columns')
    if not selected:
        selected = [col[0] for col in all_columns]
    header = [label for key, label in all_columns if key in selected]
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="customers.csv"'
    writer = csv.writer(response)
    writer.writerow(header)
    for c in customers:
        row = []
        for key, _ in all_columns:
            if key not in selected:
                continue
            row.append(getattr(c, key, ''))
        writer.writerow(row)
    return response

@login_required
def customers_import(request):
    if request.method == 'POST' and request.FILES.get('csv_file'):
        csv_file = request.FILES['csv_file']
        try:
            decoded = csv_file.read().decode('utf-8')
            reader = csv.reader(io.StringIO(decoded))
            header = next(reader, None)
            count_created = 0
            count_updated = 0
            skipped = 0
            skipped_rows = []
            for idx, row in enumerate(reader, start=2):
                # If header has 'id', expect id as first column
                if header and 'id' in [h.lower() for h in header]:
                    id_idx = [h.lower() for h in header].index('id')
                    name_idx = [h.lower() for h in header].index('name') if 'name' in [h.lower() for h in header] else 1
                    contact_idx = [h.lower() for h in header].index('contact') if 'contact' in [h.lower() for h in header] else 2
                    address_idx = [h.lower() for h in header].index('address') if 'address' in [h.lower() for h in header] else 3
                    try:
                        customer_id = row[id_idx].strip()
                        name = row[name_idx].strip()
                        contact = row[contact_idx].strip() if contact_idx < len(row) else ''
                        address = row[address_idx].strip() if address_idx < len(row) else ''
                        if customer_id:
                            try:
                                customer = Customer.objects.get(id=customer_id)
                                customer.name = name
                                customer.contact = contact
                                customer.address = address
                                customer.save()
                                count_updated += 1
                                continue
                            except Customer.DoesNotExist:
                                pass
                        if not name:
                            skipped += 1
                            skipped_rows.append(f"Row {idx}: Missing required name")
                            continue
                        Customer.objects.create(name=name, contact=contact, address=address)
                        count_created += 1
                    except Exception as e:
                        skipped += 1
                        skipped_rows.append(f"Row {idx}: {e}")
                else:
                    if len(row) < 3:
                        skipped += 1
                        skipped_rows.append(f"Row {idx}: Not enough columns")
                        continue
                    name = row[0].strip()
                    contact = row[1].strip()
                    address = row[2].strip()
                    if not name:
                        skipped += 1
                        skipped_rows.append(f"Row {idx}: Missing required name")
                        continue
                    Customer.objects.create(name=name, contact=contact, address=address)
                    count_created += 1
            msg = f"Created {count_created} customers. Updated {count_updated}."
            if skipped:
                msg += f" Skipped {skipped} row(s)."
                for reason in skipped_rows[:5]:
                    msg += f" {reason}."
                if skipped > 5:
                    msg += " ..."
            messages.info(request, msg)
        except Exception as e:
            messages.error(request, f'Error importing CSV: {e}')
        return redirect('customers')
    messages.error(request, 'No file uploaded.')
    return redirect('customers')

@login_required
def customers_export_excel(request):
    query = request.GET.get('q', '')
    customers = Customer.objects.all()
    if query:
        customers = customers.filter(
            Q(name__icontains=query) |
            Q(contact__icontains=query) |
            Q(address__icontains=query)
        )
    all_columns = [
        ('name', 'Name'),
        ('contact', 'Contact'),
        ('address', 'Address'),
        ('created_at', 'Created At'),
    ]
    selected = request.GET.getlist('columns')
    if not selected:
        selected = [col[0] for col in all_columns]
    header = [label for key, label in all_columns if key in selected]
    wb = Workbook()
    ws = wb.active
    ws.append(header)
    for c in customers:
        row = []
        for key, _ in all_columns:
            if key not in selected:
                continue
            row.append(getattr(c, key, ''))
        ws.append(row)
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="customers.xlsx"'
    wb.save(response)
    return response

# Inventory CRUD

@login_required
def inventory(request):
    query = request.GET.get('q', '')
    items = InventoryItem.objects.all()
    if query:
        items = items.filter(
            Q(item_name__icontains=query) |
            Q(fabric_type__icontains=query) |
            Q(supplier__icontains=query) |
            Q(color__icontains=query) |
            Q(size__icontains=query)
        )
    form = InventoryItemForm()
    if request.method == 'POST':
        form = InventoryItemForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('inventory')
    return render(request, 'core/inventory.html', {'items': items, 'form': form, 'query': query})

@login_required
def edit_inventory(request, pk):
    item = get_object_or_404(InventoryItem, pk=pk)
    if request.method == 'POST':
        form = InventoryItemForm(request.POST, request.FILES, instance=item)
        if form.is_valid():
            form.save()
            return redirect('inventory')
    else:
        form = InventoryItemForm(instance=item)
    return render(request, 'core/edit_inventory.html', {'form': form, 'item': item})

@permission_required('core.delete_inventoryitem', raise_exception=True)
def delete_inventory(request, pk):
    item = get_object_or_404(InventoryItem, pk=pk)
    if request.method == 'POST':
        item.delete()
        return redirect('inventory')
    return render(request, 'core/delete_inventory.html', {'item': item})

@login_required
def inventory_export(request):
    query = request.GET.get('q', '')
    items = InventoryItem.objects.all()
    if query:
        items = items.filter(
            Q(item_name__icontains=query) |
            Q(fabric_type__icontains=query) |
            Q(supplier__icontains=query) |
            Q(color__icontains=query) |
            Q(size__icontains=query)
        )
    all_columns = [
        ('item_name', 'Item Name'),
        ('item_type', 'Type'),
        ('fabric_type', 'Fabric'),
        ('cost_per_meter', 'Cost/m'),
        ('total_meters', 'Total m'),
        ('taxes', 'Taxes'),
        ('size', 'Size'),
        ('color', 'Color'),
        ('is_printed', 'Printed'),
        ('stock_quantity', 'Stock'),
        ('supplier', 'Supplier'),
    ]
    selected = request.GET.getlist('columns')
    if not selected:
        selected = [col[0] for col in all_columns]
    header = [label for key, label in all_columns if key in selected]
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="inventory.csv"'
    writer = csv.writer(response)
    writer.writerow(header)
    for i in items:
        row = []
        for key, _ in all_columns:
            if key not in selected:
                continue
            if key == 'item_type':
                row.append(i.get_item_type_display())
            elif key == 'is_printed':
                row.append('Yes' if i.is_printed else 'No')
            else:
                row.append(getattr(i, key, ''))
        writer.writerow(row)
    return response

@login_required
def inventory_import(request):
    if request.method == 'POST' and request.FILES.get('csv_file'):
        csv_file = request.FILES['csv_file']
        try:
            decoded = csv_file.read().decode('utf-8')
            reader = csv.reader(io.StringIO(decoded))
            header = next(reader, None)
            count_created = 0
            count_updated = 0
            skipped = 0
            skipped_rows = []
            for idx, row in enumerate(reader, start=2):
                # If header has 'id', expect id as first column
                if header and 'id' in [h.lower() for h in header]:
                    id_idx = [h.lower() for h in header].index('id')
                    try:
                        item_id = row[id_idx].strip()
                        if item_id:
                            try:
                                item = InventoryItem.objects.get(id=item_id)
                                # Update fields by header
                                for i, h in enumerate(header):
                                    h = h.lower()
                                    if h == 'id': continue
                                    if hasattr(item, h) and i < len(row):
                                        setattr(item, h, row[i].strip())
                                item.save()
                                count_updated += 1
                                continue
                            except InventoryItem.DoesNotExist:
                                pass
                        # If not found, create new
                        fields = [row[i].strip() if i < len(row) else '' for i in range(len(header))]
                        InventoryItem.objects.create(
                            item_name=fields[1],
                            item_type=fields[2],
                            fabric_type=fields[3],
                            cost_per_meter=fields[4],
                            total_meters=fields[5],
                            taxes=fields[6],
                            size=fields[7],
                            color=fields[8],
                            is_printed=(fields[9].lower() == 'yes'),
                            stock_quantity=fields[10],
                            supplier=fields[11]
                        )
                        count_created += 1
                    except Exception as e:
                        skipped += 1
                        skipped_rows.append(f"Row {idx}: {e}")
                else:
                    if len(row) < 11:
                        skipped += 1
                        skipped_rows.append(f"Row {idx}: Not enough columns for inventory item")
                        continue
                    InventoryItem.objects.create(
                        item_name=row[0],
                        item_type='stitched' if row[1].lower().startswith('s') else 'unstitched',
                        fabric_type=row[2],
                        cost_per_meter=row[3],
                        total_meters=row[4],
                        taxes=row[5],
                        size=row[6],
                        color=row[7],
                        is_printed=(row[8].strip().lower() == 'yes'),
                        stock_quantity=row[9],
                        supplier=row[10]
                    )
                    count_created += 1
            msg = f"Created {count_created} inventory items. Updated {count_updated}."
            if skipped:
                msg += f" Skipped {skipped} row(s)."
                for reason in skipped_rows[:5]:
                    msg += f" {reason}."
                if skipped > 5:
                    msg += " ..."
            messages.info(request, msg)
        except Exception as e:
            messages.error(request, f'Error importing CSV: {e}')
        return redirect('inventory')
    messages.error(request, 'No file uploaded.')
    return redirect('inventory')

@login_required
def inventory_export_excel(request):
    query = request.GET.get('q', '')
    items = InventoryItem.objects.all()
    if query:
        items = items.filter(
            Q(item_name__icontains=query) |
            Q(fabric_type__icontains=query) |
            Q(supplier__icontains=query) |
            Q(color__icontains=query) |
            Q(size__icontains=query)
        )
    all_columns = [
        ('item_name', 'Item Name'),
        ('item_type', 'Type'),
        ('fabric_type', 'Fabric'),
        ('cost_per_meter', 'Cost/m'),
        ('total_meters', 'Total m'),
        ('taxes', 'Taxes'),
        ('size', 'Size'),
        ('color', 'Color'),
        ('is_printed', 'Printed'),
        ('stock_quantity', 'Stock'),
        ('supplier', 'Supplier'),
    ]
    selected = request.GET.getlist('columns')
    if not selected:
        selected = [col[0] for col in all_columns]
    header = [label for key, label in all_columns if key in selected]
    wb = Workbook()
    ws = wb.active
    ws.append(header)
    for i in items:
        row = []
        for key, _ in all_columns:
            if key not in selected:
                continue
            if key == 'item_type':
                row.append(i.get_item_type_display())
            elif key == 'is_printed':
                row.append('Yes' if i.is_printed else 'No')
            else:
                row.append(getattr(i, key, ''))
        ws.append(row)
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="inventory.xlsx"'
    wb.save(response)
    return response

# Orders CRUD

@login_required
def orders(request):
    query = request.GET.get('q', '')
    orders = Order.objects.select_related('customer').all()
    if query:
        orders = orders.filter(
            Q(customer__name__icontains=query) |
            Q(status__icontains=query) |
            Q(product_type__icontains=query) |
            Q(notes__icontains=query)
        )
    form = OrderForm()
    if request.method == 'POST':
        form = OrderForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('orders')
    return render(request, 'core/orders.html', {'orders': orders, 'form': form, 'query': query})

@login_required
def edit_order(request, pk):
    order = get_object_or_404(Order, pk=pk)
    if request.method == 'POST':
        form = OrderForm(request.POST, request.FILES, instance=order)
        if form.is_valid():
            form.save()
            return redirect('orders')
    else:
        form = OrderForm(instance=order)
    return render(request, 'core/edit_order.html', {'form': form, 'order': order})

@permission_required('core.delete_order', raise_exception=True)
def delete_order(request, pk):
    order = get_object_or_404(Order, pk=pk)
    if request.method == 'POST':
        order.delete()
        return redirect('orders')
    return render(request, 'core/delete_order.html', {'order': order})

@login_required
def orders_export(request):
    query = request.GET.get('q', '')
    orders = Order.objects.select_related('customer').all()
    if query:
        orders = orders.filter(
            Q(customer__name__icontains=query) |
            Q(status__icontains=query) |
            Q(product_type__icontains=query) |
            Q(notes__icontains=query)
        )
    # Handle custom columns
    all_columns = [
        ('id', 'Order ID'),
        ('customer', 'Customer'),
        ('product_type', 'Product Type'),
        ('status', 'Status'),
        ('order_date', 'Order Date'),
        ('delivery_date', 'Delivery Date'),
        ('notes', 'Notes'),
    ]
    selected = request.GET.getlist('columns')
    if not selected:
        selected = [col[0] for col in all_columns]  # default: all columns
    # Prepare header
    header = [label for key, label in all_columns if key in selected]
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="orders.csv"'
    writer = csv.writer(response)
    writer.writerow(header)
    for o in orders:
        row = []
        for key, _ in all_columns:
            if key not in selected:
                continue
            if key == 'id':
                row.append(o.id)
            elif key == 'customer':
                row.append(o.customer.name)
            elif key == 'product_type':
                row.append(o.get_product_type_display())
            else:
                row.append(getattr(o, key, ''))
        writer.writerow(row)
    return response

@login_required
def orders_export_excel(request):
    query = request.GET.get('q', '')
    orders = Order.objects.select_related('customer').all()
    if query:
        orders = orders.filter(
            Q(customer__name__icontains=query) |
            Q(status__icontains=query) |
            Q(product_type__icontains=query) |
            Q(notes__icontains=query)
        )
    all_columns = [
        ('id', 'Order ID'),
        ('customer', 'Customer'),
        ('product_type', 'Product Type'),
        ('status', 'Status'),
        ('order_date', 'Order Date'),
        ('delivery_date', 'Delivery Date'),
        ('notes', 'Notes'),
    ]
    selected = request.GET.getlist('columns')
    if not selected:
        selected = [col[0] for col in all_columns]
    header = [label for key, label in all_columns if key in selected]
    wb = Workbook()
    ws = wb.active
    ws.append(header)
    for o in orders:
        row = []
        for key, _ in all_columns:
            if key not in selected:
                continue
            if key == 'id':
                row.append(o.id)
            elif key == 'customer':
                row.append(o.customer.name)
            elif key == 'product_type':
                row.append(o.get_product_type_display())
            else:
                row.append(getattr(o, key, ''))
        ws.append(row)
    # Auto-size columns
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="orders.xlsx"'
    wb.save(response)
    return response

@login_required
def orders_import(request):
    if request.method == 'POST' and request.FILES.get('csv_file'):
        csv_file = request.FILES['csv_file']
        try:
            decoded = csv_file.read().decode('utf-8')
            reader = csv.reader(io.StringIO(decoded))
            header = next(reader, None)
            count_created = 0
            count_updated = 0
            skipped = 0
            skipped_rows = []
            for idx, row in enumerate(reader, start=2):
                if header and 'id' in [h.lower() for h in header]:
                    id_idx = [h.lower() for h in header].index('id')
                    try:
                        order_id = row[id_idx].strip()
                        if order_id:
                            try:
                                order = Order.objects.get(id=order_id)
                                # Update fields by header
                                for i, h in enumerate(header):
                                    h = h.lower()
                                    if h == 'id': continue
                                    if hasattr(order, h) and i < len(row):
                                        setattr(order, h, row[i].strip())
                                order.save()
                                count_updated += 1
                                continue
                            except Order.DoesNotExist:
                                pass
                        # If not found, create new
                        customer_name = row[1].strip() if len(row) > 1 else ''
                        product_type = row[2].strip().lower() if len(row) > 2 else ''
                        status = row[3].strip() if len(row) > 3 else ''
                        order_date = row[4].strip() if len(row) > 4 else ''
                        delivery_date = row[5].strip() if len(row) > 5 else ''
                        notes = row[6].strip() if len(row) > 6 else ''
                        if not customer_name or not product_type or not status or not order_date:
                            skipped += 1
                            skipped_rows.append(f"Row {idx}: Missing required fields")
                            continue
                        customer, _ = Customer.objects.get_or_create(name=customer_name)
                        Order.objects.create(
                            customer=customer,
                            product_type=product_type,
                            status=status,
                            order_date=order_date,
                            delivery_date=delivery_date,
                            notes=notes
                        )
                        count_created += 1
                    except Exception as e:
                        skipped += 1
                        skipped_rows.append(f"Row {idx}: {e}")
                else:
                    if len(row) < 6:
                        skipped += 1
                        skipped_rows.append(f"Row {idx}: Not enough columns")
                        continue
                    customer_name = row[0].strip()
                    product_type = row[1].strip().lower()
                    status = row[2].strip()
                    order_date = row[3].strip()
                    delivery_date = row[4].strip()
                    notes = row[5].strip()
                    if not customer_name or not product_type or not status or not order_date:
                        skipped += 1
                        skipped_rows.append(f"Row {idx}: Missing required fields")
                        continue
                    customer, _ = Customer.objects.get_or_create(name=customer_name)
                    Order.objects.create(
                        customer=customer,
                        product_type=product_type,
                        status=status,
                        order_date=order_date,
                        delivery_date=delivery_date,
                        notes=notes
                    )
                    count_created += 1
            msg = f"Created {count_created} orders. Updated {count_updated}."
            if skipped:
                msg += f" Skipped {skipped} row(s)."
                for reason in skipped_rows[:5]:
                    msg += f" {reason}."
                if skipped > 5:
                    msg += " ..."
            messages.info(request, msg)
        except Exception as e:
            messages.error(request, f'Error importing CSV: {e}')
        return redirect('orders')
    messages.error(request, 'No file uploaded.')
    return redirect('orders')

# Requirements CRUD

@login_required
def requirements(request):
    query = request.GET.get('q', '')
    requirements = Requirement.objects.select_related('order').all()
    if query:
        requirements = requirements.filter(
            Q(description__icontains=query) |
            Q(order__id__icontains=query) |
            Q(notes__icontains=query) |
            Q(is_fulfilled__icontains=query)
        )
    form = RequirementForm()
    if request.method == 'POST':
        form = RequirementForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('requirements')
    return render(request, 'core/requirements.html', {'requirements': requirements, 'form': form, 'query': query})

@login_required
def edit_requirement(request, pk):
    requirement = get_object_or_404(Requirement, pk=pk)
    if request.method == 'POST':
        form = RequirementForm(request.POST, request.FILES, instance=requirement)
        if form.is_valid():
            form.save()
            return redirect('requirements')
    else:
        form = RequirementForm(instance=requirement)
    return render(request, 'core/edit_requirement.html', {'form': form, 'requirement': requirement})

@permission_required('core.delete_requirement', raise_exception=True)
def delete_requirement(request, pk):
    requirement = get_object_or_404(Requirement, pk=pk)
    if request.method == 'POST':
        requirement.delete()
        return redirect('requirements')
    return render(request, 'core/delete_requirement.html', {'requirement': requirement})

@login_required
def requirements_export(request):
    query = request.GET.get('q', '')
    requirements = Requirement.objects.select_related('order').all()
    if query:
        requirements = requirements.filter(
            Q(description__icontains=query) |
            Q(order__id__icontains=query) |
            Q(notes__icontains=query) |
            Q(is_fulfilled__icontains=query)
        )
    all_columns = [
        ('order', 'Order'),
        ('description', 'Description'),
        ('is_fulfilled', 'Fulfilled'),
        ('steps_done', 'Steps Done'),
        ('steps_not_done', 'Steps Not Done'),
        ('notes', 'Notes'),
    ]
    selected = request.GET.getlist('columns')
    if not selected:
        selected = [col[0] for col in all_columns]
    header = [label for key, label in all_columns if key in selected]
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="requirements.csv"'
    writer = csv.writer(response)
    writer.writerow(header)
    for r in requirements:
        row = []
        for key, _ in all_columns:
            if key not in selected:
                continue
            if key == 'order':
                row.append(r.order.id if r.order else '')
            elif key == 'is_fulfilled':
                row.append('Yes' if r.is_fulfilled else 'No')
            elif key == 'steps_done':
                row.append('; '.join(r.steps_done or []))
            elif key == 'steps_not_done':
                row.append('; '.join(r.steps_not_done or []))
            else:
                row.append(getattr(r, key, ''))
        writer.writerow(row)
    return response

@login_required
def requirements_import(request):
    if request.method == 'POST' and request.FILES.get('csv_file'):
        csv_file = request.FILES['csv_file']
        try:
            decoded = csv_file.read().decode('utf-8')
            reader = csv.reader(io.StringIO(decoded))
            header = next(reader, None)
            count_created = 0
            count_updated = 0
            skipped = 0
            skipped_rows = []
            for idx, row in enumerate(reader, start=2):
                if header and 'id' in [h.lower() for h in header]:
                    id_idx = [h.lower() for h in header].index('id')
                    try:
                        req_id = row[id_idx].strip()
                        if req_id:
                            try:
                                req = Requirement.objects.get(id=req_id)
                                for i, h in enumerate(header):
                                    h = h.lower()
                                    if h == 'id': continue
                                    if hasattr(req, h) and i < len(row):
                                        setattr(req, h, row[i].strip())
                                req.save()
                                count_updated += 1
                                continue
                            except Requirement.DoesNotExist:
                                pass
                        # If not found, create new
                        order_id = row[1].strip() if len(row) > 1 else ''
                        description = row[2].strip() if len(row) > 2 else ''
                        is_fulfilled = row[3].strip().lower() in ['yes', 'true', '1'] if len(row) > 3 else False
                        steps_done = [s.strip() for s in row[4].split(';') if s.strip()] if len(row) > 4 else []
                        steps_not_done = [s.strip() for s in row[5].split(';') if s.strip()] if len(row) > 5 else []
                        notes = row[6].strip() if len(row) > 6 else ''
                        if not order_id or not description:
                            skipped += 1
                            skipped_rows.append(f"Row {idx}: Missing required fields")
                            continue
                        try:
                            order = Order.objects.get(id=order_id)
                        except Order.DoesNotExist:
                            skipped += 1
                            skipped_rows.append(f"Row {idx}: Order not found")
                            continue
                        Requirement.objects.create(
                            order=order,
                            description=description,
                            is_fulfilled=is_fulfilled,
                            steps_done=steps_done,
                            steps_not_done=steps_not_done,
                            notes=notes
                        )
                        count_created += 1
                    except Exception as e:
                        skipped += 1
                        skipped_rows.append(f"Row {idx}: {e}")
                else:
                    if len(row) < 6:
                        skipped += 1
                        skipped_rows.append(f"Row {idx}: Not enough columns")
                        continue
                    order_id = row[0].strip()
                    description = row[1].strip()
                    is_fulfilled = row[2].strip().lower() in ['yes', 'true', '1']
                    steps_done = [s.strip() for s in row[3].split(';') if s.strip()]
                    steps_not_done = [s.strip() for s in row[4].split(';') if s.strip()]
                    notes = row[5].strip()
                    if not order_id or not description:
                        skipped += 1
                        skipped_rows.append(f"Row {idx}: Missing required fields")
                        continue
                    try:
                        order = Order.objects.get(id=order_id)
                    except Order.DoesNotExist:
                        skipped += 1
                        skipped_rows.append(f"Row {idx}: Order not found")
                        continue
                    Requirement.objects.create(
                        order=order,
                        description=description,
                        is_fulfilled=is_fulfilled,
                        steps_done=steps_done,
                        steps_not_done=steps_not_done,
                        notes=notes
                    )
                    count_created += 1
            msg = f"Created {count_created} requirements. Updated {count_updated}."
            if skipped:
                msg += f" Skipped {skipped} row(s)."
                for reason in skipped_rows[:5]:
                    msg += f" {reason}."
                if skipped > 5:
                    msg += " ..."
            messages.info(request, msg)
        except Exception as e:
            messages.error(request, f'Error importing CSV: {e}')
        return redirect('requirements')
    messages.error(request, 'No file uploaded.')
    return redirect('requirements')

@login_required
def requirements_export_excel(request):
    query = request.GET.get('q', '')
    requirements = Requirement.objects.select_related('order').all()
    if query:
        requirements = requirements.filter(
            Q(description__icontains=query) |
            Q(order__id__icontains=query) |
            Q(notes__icontains=query) |
            Q(is_fulfilled__icontains=query)
        )
    all_columns = [
        ('order', 'Order'),
        ('description', 'Description'),
        ('is_fulfilled', 'Fulfilled'),
        ('steps_done', 'Steps Done'),
        ('steps_not_done', 'Steps Not Done'),
        ('notes', 'Notes'),
    ]
    selected = request.GET.getlist('columns')
    if not selected:
        selected = [col[0] for col in all_columns]
    header = [label for key, label in all_columns if key in selected]
    wb = Workbook()
    ws = wb.active
    ws.append(header)
    for r in requirements:
        row = []
        for key, _ in all_columns:
            if key not in selected:
                continue
            if key == 'order':
                row.append(r.order.id if r.order else '')
            elif key == 'is_fulfilled':
                row.append('Yes' if r.is_fulfilled else 'No')
            elif key == 'steps_done':
                row.append('; '.join(r.steps_done or []))
            elif key == 'steps_not_done':
                row.append('; '.join(r.steps_not_done or []))
            else:
                row.append(getattr(r, key, ''))
        ws.append(row)
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="requirements.xlsx"'
    wb.save(response)
    return response

# Payments CRUD

@login_required
def payments(request):
    query = request.GET.get('q', '')
    payments = Payment.objects.select_related('order').all()
    if query:
        payments = payments.filter(
            Q(order__id__icontains=query) |
            Q(status__icontains=query) |
            Q(notes__icontains=query) |
            Q(amount__icontains=query)
        )
    form = PaymentForm()
    if request.method == 'POST':
        form = PaymentForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('payments')
    return render(request, 'core/payments.html', {'payments': payments, 'form': form, 'query': query})

@login_required
def edit_payment(request, pk):
    payment = get_object_or_404(Payment, pk=pk)
    if request.method == 'POST':
        form = PaymentForm(request.POST, instance=payment)
        if form.is_valid():
            form.save()
            return redirect('payments')
    else:
        form = PaymentForm(instance=payment)
    return render(request, 'core/edit_payment.html', {'form': form, 'payment': payment})

@permission_required('core.delete_payment', raise_exception=True)
def delete_payment(request, pk):
    payment = get_object_or_404(Payment, pk=pk)
    if request.method == 'POST':
        payment.delete()
        return redirect('payments')
    return render(request, 'core/delete_payment.html', {'payment': payment})

@login_required
def payments_export(request):
    query = request.GET.get('q', '')
    payments = Payment.objects.select_related('order').all()
    if query:
        payments = payments.filter(
            Q(order__id__icontains=query) |
            Q(status__icontains=query) |
            Q(notes__icontains=query) |
            Q(amount__icontains=query)
        )
    all_columns = [
        ('order', 'Order'),
        ('amount', 'Amount'),
        ('status', 'Status'),
        ('payment_date', 'Payment Date'),
        ('notes', 'Notes'),
    ]
    selected = request.GET.getlist('columns')
    if not selected:
        selected = [col[0] for col in all_columns]
    header = [label for key, label in all_columns if key in selected]
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="payments.csv"'
    writer = csv.writer(response)
    writer.writerow(header)
    for p in payments:
        row = []
        for key, _ in all_columns:
            if key not in selected:
                continue
            if key == 'order':
                row.append(p.order.id if p.order else '')
            else:
                row.append(getattr(p, key, ''))
        writer.writerow(row)
    return response

@login_required
def payments_import(request):
    if request.method == 'POST' and request.FILES.get('csv_file'):
        csv_file = request.FILES['csv_file']
        try:
            decoded = csv_file.read().decode('utf-8')
            reader = csv.reader(io.StringIO(decoded))
            header = next(reader, None)
            count_created = 0
            count_updated = 0
            skipped = 0
            skipped_rows = []
            for idx, row in enumerate(reader, start=2):
                if header and 'id' in [h.lower() for h in header]:
                    id_idx = [h.lower() for h in header].index('id')
                    try:
                        payment_id = row[id_idx].strip()
                        if payment_id:
                            try:
                                payment = Payment.objects.get(id=payment_id)
                                for i, h in enumerate(header):
                                    h = h.lower()
                                    if h == 'id': continue
                                    if hasattr(payment, h) and i < len(row):
                                        setattr(payment, h, row[i].strip())
                                payment.save()
                                count_updated += 1
                                continue
                            except Payment.DoesNotExist:
                                pass
                        # If not found, create new
                        order_id = row[1].strip() if len(row) > 1 else ''
                        amount = row[2].strip() if len(row) > 2 else ''
                        status = row[3].strip() if len(row) > 3 else ''
                        payment_date = row[4].strip() if len(row) > 4 else ''
                        notes = row[5].strip() if len(row) > 5 else ''
                        if not order_id or not amount or not status:
                            skipped += 1
                            skipped_rows.append(f"Row {idx}: Missing required fields")
                            continue
                        try:
                            order = Order.objects.get(id=order_id)
                        except Order.DoesNotExist:
                            skipped += 1
                            skipped_rows.append(f"Row {idx}: Order not found")
                            continue
                        Payment.objects.create(
                            order=order,
                            amount=amount,
                            status=status,
                            payment_date=payment_date if payment_date else None,
                            notes=notes
                        )
                        count_created += 1
                    except Exception as e:
                        skipped += 1
                        skipped_rows.append(f"Row {idx}: {e}")
                else:
                    if len(row) < 5:
                        skipped += 1
                        skipped_rows.append(f"Row {idx}: Not enough columns")
                        continue
                    order_id = row[0].strip()
                    amount = row[1].strip()
                    status = row[2].strip()
                    payment_date = row[3].strip()
                    notes = row[4].strip()
                    if not order_id or not amount or not status:
                        skipped += 1
                        skipped_rows.append(f"Row {idx}: Missing required fields")
                        continue
                    try:
                        order = Order.objects.get(id=order_id)
                    except Order.DoesNotExist:
                        skipped += 1
                        skipped_rows.append(f"Row {idx}: Order not found")
                        continue
                    Payment.objects.create(
                        order=order,
                        amount=amount,
                        status=status,
                        payment_date=payment_date if payment_date else None,
                        notes=notes
                    )
                    count_created += 1
            msg = f"Created {count_created} payments. Updated {count_updated}."
            if skipped:
                msg += f" Skipped {skipped} row(s)."
                for reason in skipped_rows[:5]:
                    msg += f" {reason}."
                if skipped > 5:
                    msg += " ..."
            messages.info(request, msg)
        except Exception as e:
            messages.error(request, f'Error importing CSV: {e}')
        return redirect('payments')
    messages.error(request, 'No file uploaded.')
    return redirect('payments')

@login_required
def payments_export_excel(request):
    query = request.GET.get('q', '')
    payments = Payment.objects.select_related('order').all()
    if query:
        payments = payments.filter(
            Q(order__id__icontains=query) |
            Q(status__icontains=query) |
            Q(notes__icontains=query) |
            Q(amount__icontains=query)
        )
    all_columns = [
        ('order', 'Order'),
        ('amount', 'Amount'),
        ('status', 'Status'),
        ('payment_date', 'Payment Date'),
        ('notes', 'Notes'),
    ]
    selected = request.GET.getlist('columns')
    if not selected:
        selected = [col[0] for col in all_columns]
    header = [label for key, label in all_columns if key in selected]
    wb = Workbook()
    ws = wb.active
    ws.append(header)
    for p in payments:
        row = []
        for key, _ in all_columns:
            if key not in selected:
                continue
            if key == 'order':
                row.append(p.order.id if p.order else '')
            else:
                row.append(getattr(p, key, ''))
        ws.append(row)
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="payments.xlsx"'
    wb.save(response)
    return response

# Suppliers CRUD

@login_required
def suppliers(request):
    query = request.GET.get('q', '')
    suppliers = Supplier.objects.all()
    if query:
        suppliers = suppliers.filter(
            Q(name__icontains=query) |
            Q(contact__icontains=query) |
            Q(address__icontains=query) |
            Q(email__icontains=query) |
            Q(phone__icontains=query)
        )
    form = SupplierForm()
    if request.method == 'POST':
        form = SupplierForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('suppliers')
    return render(request, 'core/suppliers.html', {'suppliers': suppliers, 'form': form, 'query': query})

@login_required
def edit_supplier(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)
    if request.method == 'POST':
        form = SupplierForm(request.POST, instance=supplier)
        if form.is_valid():
            form.save()
            return redirect('suppliers')
    else:
        form = SupplierForm(instance=supplier)
    return render(request, 'core/edit_supplier.html', {'form': form, 'supplier': supplier})

@permission_required('core.delete_supplier', raise_exception=True)
def delete_supplier(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)
    if request.method == 'POST':
        supplier.delete()
        return redirect('suppliers')
    return render(request, 'core/delete_supplier.html', {'supplier': supplier})

# Purchases CRUD

@login_required
def purchases(request):
    query = request.GET.get('q', '')
    purchases = Purchase.objects.select_related('supplier', 'item').all()
    if query:
        purchases = purchases.filter(
            Q(supplier__name__icontains=query) |
            Q(item__item_name__icontains=query) |
            Q(notes__icontains=query)
        )
    form = PurchaseForm()
    if request.method == 'POST':
        form = PurchaseForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('purchases')
    return render(request, 'core/purchases.html', {'purchases': purchases, 'form': form, 'query': query})

@login_required
def edit_purchase(request, pk):
    purchase = get_object_or_404(Purchase, pk=pk)
    if request.method == 'POST':
        form = PurchaseForm(request.POST, instance=purchase)
        if form.is_valid():
            form.save()
            return redirect('purchases')
    else:
        form = PurchaseForm(instance=purchase)
    return render(request, 'core/edit_purchase.html', {'form': form, 'purchase': purchase})

@permission_required('core.delete_purchase', raise_exception=True)
def delete_purchase(request, pk):
    purchase = get_object_or_404(Purchase, pk=pk)
    if request.method == 'POST':
        purchase.delete()
        return redirect('purchases')
    return render(request, 'core/delete_purchase.html', {'purchase': purchase})

# Meeting Mode

@login_required
def meeting_mode(request):
    customer_form = CustomerForm(prefix='customer')
    order_form = OrderForm(prefix='order')
    requirement_form = RequirementForm(prefix='requirement')
    payment_form = PaymentForm(prefix='payment')
    all_forms = [customer_form, order_form, requirement_form, payment_form]
    success = False

    if request.method == 'POST':
        customer_form = CustomerForm(request.POST, prefix='customer')
        order_form = OrderForm(request.POST, request.FILES, prefix='order')
        requirement_form = RequirementForm(request.POST, prefix='requirement')
        payment_form = PaymentForm(request.POST, prefix='payment')
        all_forms = [customer_form, order_form, requirement_form, payment_form]

        if customer_form.is_valid():
            customer = customer_form.save()
            if order_form.is_valid():
                order = order_form.save(commit=False)
                order.customer = customer
                order.save()
                if requirement_form.is_valid():
                    requirement = requirement_form.save(commit=False)
                    requirement.order = order
                    requirement.save()
                if payment_form.is_valid():
                    payment = payment_form.save(commit=False)
                    payment.order = order
                    payment.save()
                success = True
                return redirect('dashboard')

    return render(request, 'core/meeting_mode.html', {
        'customer_form': customer_form,
        'order_form': order_form,
        'requirement_form': requirement_form,
        'payment_form': payment_form,
        'all_forms': all_forms,
        'success': success,
    })

# Dashboard

@login_required
def dashboard(request):
    total_customers = Customer.objects.count()
    total_orders = Order.objects.count()
    pending_orders = Order.objects.filter(status__iexact='Pending').count()
    total_inventory = InventoryItem.objects.count()
    low_stock_items = InventoryItem.objects.filter(stock_quantity__lte=5)
    total_payments = Payment.objects.aggregate(total=Sum('amount'))['total'] or 0
    outstanding_payments = Payment.objects.filter(status__iexact='Pending').aggregate(total=Sum('amount'))['total'] or 0
    total_suppliers = Supplier.objects.count()
    total_purchases = Purchase.objects.count()

    return render(request, 'core/dashboard.html', {
        'total_customers': total_customers,
        'total_orders': total_orders,
        'pending_orders': pending_orders,
        'total_inventory': total_inventory,
        'low_stock_items': low_stock_items,
        'total_payments': total_payments,
        'outstanding_payments': outstanding_payments,
        'total_suppliers': total_suppliers,
        'total_purchases': total_purchases,
    })

@login_required
def analytics(request):
    # Orders per month
    orders_by_month = (
        Order.objects.annotate(month=TruncMonth('order_date'))
        .values('month')
        .annotate(count=Count('id'))
        .order_by('month')
    )
    # Revenue per month (sum of payments)
    revenue_by_month = (
        Payment.objects.annotate(month=TruncMonth('payment_date'))
        .values('month')
        .annotate(total=Sum('amount'))
        .order_by('month')
    )
    # Inventory stock by item
    inventory_stock = (
        InventoryItem.objects.values('item_name').annotate(stock=Sum('stock_quantity')).order_by('-stock')[:10]
    )
    # Top customers by order count
    top_customers = (
        Customer.objects.annotate(order_count=Count('orders')).order_by('-order_count')[:10]
    )
    return render(request, 'core/analytics.html', {
        'orders_by_month': list(orders_by_month),
        'revenue_by_month': list(revenue_by_month),
        'inventory_stock': list(inventory_stock),
        'top_customers': list(top_customers),
    })

@login_required
def model_history(request, model_name, object_id):
    model = apps.get_model('core', model_name)
    obj = model.objects.get(pk=object_id)
    history = obj.history.all().order_by('-history_date')
    return render(request, 'core/model_history.html', {'object': obj, 'history': history, 'model_name': model_name})

def register(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Registration successful. You can now log in.')
            return redirect('/accounts/login/')
    else:
        form = UserCreationForm()
    return render(request, 'registration/register.html', {'form': form})

# Notification views
@login_required
def notifications(request):
    notes = Notification.objects.filter(user=request.user).order_by('-created_at')
    return render(request, 'core/notifications.html', {'notifications': notes})
@login_required
def mark_notification_read(request, pk):
    note = Notification.objects.get(pk=pk, user=request.user)
    note.is_read = True
    note.save()
    return redirect('notifications')

# Customer portal views
class CustomerLoginView(LoginView):
    template_name = 'core/customer_login.html'
    def get_success_url(self):
        return '/customer/dashboard/'
@login_required
def customer_dashboard(request):
    cu = CustomerUser.objects.get(user=request.user)
    orders = Order.objects.filter(customer=cu.customer)
    return render(request, 'core/customer_dashboard.html', {'orders': orders})
@login_required
def customer_invoice_pdf(request, order_id):
    order = Order.objects.get(pk=order_id)
    buf = io.BytesIO()
    p = canvas.Canvas(buf, pagesize=letter)
    p.drawString(100, 750, f"Invoice for Order #{order.id}")
    p.drawString(100, 730, f"Customer: {order.customer.name}")
    p.drawString(100, 710, f"Status: {order.status}")
    p.drawString(100, 690, f"Total: ...")
    p.showPage()
    p.save()
    buf.seek(0)
    return FileResponse(buf, as_attachment=True, filename=f'invoice_order_{order.id}.pdf')

# Barcode/QR code generation
@login_required
def order_qrcode(request, order_id):
    order = Order.objects.get(pk=order_id)
    qr = qrcode.make(f"Order ID: {order.id}")
    buf = io.BytesIO()
    qr.save(buf, format='PNG')
    buf.seek(0)
    return FileResponse(buf, content_type='image/png')
@login_required
def inventory_qrcode(request, item_id):
    item = InventoryItem.objects.get(pk=item_id)
    qr = qrcode.make(f"Item: {item.item_name}")
    buf = io.BytesIO()
    qr.save(buf, format='PNG')
    buf.seek(0)
    return FileResponse(buf, content_type='image/png')

# Calendar view for orders
@login_required
def orders_calendar(request):
    orders = Order.objects.all()
    events = [
        {"title": f"Order #{o.id}", "start": str(o.order_date), "end": str(o.delivery_date) if o.delivery_date else str(o.order_date)}
        for o in orders
    ]
    return render(request, 'core/orders_calendar.html', {'events': events})

# iCal export for orders
@login_required
def orders_ical(request):
    cal = Calendar()
    for o in Order.objects.all():
        event = Event()
        event.add('summary', f'Order #{o.id}')
        event.add('dtstart', o.order_date)
        if o.delivery_date:
            event.add('dtend', o.delivery_date)
        cal.add_component(event)
    response = HttpResponse(cal.to_ical(), content_type='text/calendar')
    response['Content-Disposition'] = 'attachment; filename="orders.ics"'
    return response

# Email notification utility
from django.conf import settings
def send_order_status_email(order):
    send_mail(
        f"Order #{order.id} Status Update",
        f"Your order status is now: {order.status}",
        settings.DEFAULT_FROM_EMAIL,
        [order.customer.email],
        fail_silently=True,
    )

# Celery task stubs (for demo, not actual celery integration)
def schedule_payment_reminder(order_id):
    # Would be a celery task in production
    order = Order.objects.get(pk=order_id)
    send_mail(
        f"Payment Reminder for Order #{order.id}",
        "Please complete your payment.",
        settings.DEFAULT_FROM_EMAIL,
        [order.customer.email],
        fail_silently=True,
    )

@login_required
def sample_customers_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="sample_customers.csv"'
    writer = csv.writer(response)
    writer.writerow(['id', 'name', 'contact', 'address'])
    writer.writerow(['# id: leave blank to create new, or set to update existing'])
    writer.writerow(['', 'John Doe', '1234567890', '123 Main St'])
    return response

@login_required
def sample_inventory_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="sample_inventory.csv"'
    writer = csv.writer(response)
    writer.writerow(['id', 'item_name', 'item_type', 'fabric_type', 'cost_per_meter', 'total_meters', 'taxes', 'size', 'color', 'is_printed', 'stock_quantity', 'supplier'])
    writer.writerow(['# id: leave blank to create new, or set to update existing'])
    writer.writerow(['', 'Cotton Roll', 'unstitched', 'Cotton', '100', '50', '5', 'L', 'White', 'False', '100', 'ABC Supplier'])
    return response

@login_required
def sample_orders_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="sample_orders.csv"'
    writer = csv.writer(response)
    writer.writerow(['id', 'customer', 'inventory_item', 'product_type', 'measurements', 'status', 'notes', 'delivery_date'])
    writer.writerow(['# id: leave blank to create new, or set to update existing'])
    writer.writerow(['', 'John Doe', 'Cotton Roll', 'stitched', '{"length": 40, "chest": 36}', 'Pending', 'Urgent', '2024-08-01'])
    return response

@login_required
def sample_requirements_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="sample_requirements.csv"'
    writer = csv.writer(response)
    writer.writerow(['id', 'order', 'description', 'is_fulfilled', 'steps_done', 'steps_not_done', 'notes'])
    writer.writerow(['# id: leave blank to create new, or set to update existing'])
    writer.writerow(['', '1', 'Cutting and stitching', 'False', '["Cutting"]', '["Stitching"]', ''])
    return response

@login_required
def sample_payments_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="sample_payments.csv"'
    writer = csv.writer(response)
    writer.writerow(['id', 'order', 'amount', 'status', 'payment_date', 'notes'])
    writer.writerow(['# id: leave blank to create new, or set to update existing'])
    writer.writerow(['', '1', '500', 'Pending', '2024-08-01', 'Advance'])
    return response
